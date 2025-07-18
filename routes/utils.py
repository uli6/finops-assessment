from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, send_file
import sqlite3
import json
import markdown
import html
from datetime import datetime
from models.database import get_db_connection
from services.ai_service import evaluate_finops_maturity
from config import DATABASE
from functools import wraps

utils_bp = Blueprint('utils', __name__)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('auth.index'))
        return f(*args, **kwargs)
    return decorated_function


def get_user_by_email(email):
    """Fetch a user row by email (normalized and hashed)"""
    from services.encryption_service import encryption_service
    from routes.auth import normalize_email
    email = normalize_email(email.strip().lower())
    email_hash = encryption_service.hash_email(email)
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE email_hash = ?', (email_hash,))
        return cursor.fetchone()

def get_maturity_label(avg):
    """Map average score to maturity label ('Crawl', 'Walk', 'Run')"""
    if avg is None:
        return 'N/A'
    if avg < 1.5:
        return 'Crawl'
    elif avg < 3:
        return 'Walk'
    else:
        return 'Run'

@utils_bp.route('/test_openai')
def test_openai():
    """Test OpenAI API connection"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        # Test with a simple evaluation
        result = evaluate_finops_maturity("Data Ingestion", "Knowledge", "Good understanding", "", "complete")
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@utils_bp.route('/reprocess_response/<int:assessment_id>/<capability_id>/<lens_id>')
def reprocess_response(assessment_id, capability_id, lens_id):
    """Reprocess a specific response with AI"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verify assessment belongs to user
    cursor.execute('SELECT user_id FROM assessments WHERE id = ?', (assessment_id,))
    assessment = cursor.fetchone()
    
    if not assessment or assessment[0] != session['user_id']:
        conn.close()
        return jsonify({'error': 'Assessment not found'}), 404
    
    # Get the response
    cursor.execute('''
        SELECT answer FROM responses 
        WHERE assessment_id = ? AND capability_id = ? AND lens_id = ?
    ''', (assessment_id, capability_id, lens_id))
    response = cursor.fetchone()
    
    if not response:
        conn.close()
        return jsonify({'error': 'Response not found'}), 404
    
    answer = response[0]
    
    # Get capability and lens names
    from data.capabilities import CAPABILITIES, LENSES
    
    capability = next((c for c in CAPABILITIES if c['id'] == capability_id), None)
    lens = next((l for l in LENSES if l['id'] == lens_id), None)
    
    if not capability or not lens:
        conn.close()
        return jsonify({'error': 'Invalid capability or lens'}), 400
    
    try:
        # Reprocess with AI
        evaluation = evaluate_finops_maturity(capability['name'], lens['name'], answer, '', 'complete')
        score = evaluation['score']
        improvement_suggestions = evaluation['improvement_suggestions']
        
        # Update the response
        cursor.execute('''
            UPDATE responses 
            SET score = ?, improvement_suggestions = ?
            WHERE assessment_id = ? AND capability_id = ? AND lens_id = ?
        ''', (score, improvement_suggestions, assessment_id, capability_id, lens_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'score': score, 
            'improvement_suggestions': improvement_suggestions
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Failed to reprocess: {str(e)}'}), 500

@utils_bp.route('/my_responses')
def my_responses():
    """View all user responses"""
    if 'user_id' not in session:
        return redirect(url_for('auth.index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all user's responses with assessment details
    cursor.execute('''
        SELECT 
            r.assessment_id,
            a.domain,
            r.capability_id,
            r.lens_id,
            r.answer,
            r.score,
            r.improvement_suggestions,
            r.created_at
        FROM responses r
        JOIN assessments a ON r.assessment_id = a.id
        WHERE a.user_id = ?
        ORDER BY r.created_at DESC
    ''', (session['user_id'],))
    responses = cursor.fetchall()
    
    conn.close()
    
    return render_template('my_responses.html', responses=responses)

@utils_bp.route('/settings')
def settings():
    """User settings page"""
    if 'user_id' not in session:
        return redirect(url_for('auth.index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT created_at, company_role FROM users WHERE id = ?', (session['user_id'],))
    user = cursor.fetchone()
    conn.close()
    
    user_created_at = user[0][:10] if user and user[0] else 'N/A'
    company_role = user[1] if user and user[1] else 'N/A'
    
    return render_template('settings.html', user_created_at=user_created_at, company_role=company_role)

@utils_bp.route('/delete_account', methods=['DELETE'])
def delete_account():
    """Delete user account and all data"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Delete all user data
        cursor.execute('DELETE FROM responses WHERE assessment_id IN (SELECT id FROM assessments WHERE user_id = ?)', (session['user_id'],))
        cursor.execute('DELETE FROM assessments WHERE user_id = ?', (session['user_id'],))
        cursor.execute('DELETE FROM users WHERE id = ?', (session['user_id'],))
        
        conn.commit()
        conn.close()
        
        session.clear()
        return jsonify({'success': True, 'message': 'Account deleted successfully'})
        
    except Exception as e:
        conn.close()
        return jsonify({'error': f'Failed to delete account: {str(e)}'}), 500

@utils_bp.route('/export_pdf/<int:assessment_id>')
def export_pdf(assessment_id):
    """Export assessment results as PDF"""
    if 'user_id' not in session:
        return redirect(url_for('auth.index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get assessment details
    cursor.execute('''
        SELECT scope_id, domain, status, overall_percentage, recommendations, created_at, updated_at
        FROM assessments 
        WHERE id = ? AND user_id = ?
    ''', (assessment_id, session['user_id']))
    assessment = cursor.fetchone()
    
    if not assessment:
        conn.close()
        return render_template('dashboard.html', error='Assessment not found')
    
    scope_id, domain, status, overall_percentage, recommendations, created_at, updated_at = assessment
    
    # Get responses (only the latest response per capability_id + lens_id combination)
    cursor.execute('''
        SELECT capability_id, lens_id, answer, score, improvement_suggestions
        FROM responses r1
        WHERE assessment_id = ? 
        AND id = (
            SELECT MAX(id) 
            FROM responses r2 
            WHERE r2.assessment_id = r1.assessment_id 
            AND r2.capability_id = r1.capability_id 
            AND r2.lens_id = r1.lens_id
        )
        ORDER BY capability_id, lens_id
    ''', (assessment_id,))
    responses = cursor.fetchall()
    
    # Calculate total score and possible points
    if domain == 'Complete Assessment':
        from data.capabilities import CAPABILITIES, LENSES
        domain_capabilities = CAPABILITIES
        lenses = LENSES
        total_questions = 105
        total_possible_points = 105 * 4
    else:
        from data.capabilities import CAPABILITIES, LENSES
        domain_capabilities = [cap for cap in CAPABILITIES if cap['domain'] == domain]
        lenses = LENSES
        total_questions = len(domain_capabilities) * 5
        total_possible_points = total_questions * 4
    total_score = sum(response[3] for response in responses) if responses else 0
    raw_average = (total_score / total_questions) if total_questions > 0 else 0
    
    user_maturity_label = get_maturity_label(raw_average)
    # Calculate correct overall percentage
    correct_overall_percentage = (total_score / total_possible_points) * 100 if total_possible_points > 0 else 0
    
    # Calculate unique questions answered
    unique_questions_answered = len(responses) if responses else 0
    
    # Calculate domain-specific benchmarks and scores (mimic results page logic)
    # ... (copy logic from get_assessment_results)
    # For brevity, only include basic industry_avg and user_score for now
    industry_avg = 0
    user_score = 0
    has_benchmark_data = False
    domain_benchmarks = None
    domain_scores = None
    
    # Parse recommendations
    def parse_recommendations(recommendations):
        if not recommendations:
            return []
        try:
            parsed = json.loads(recommendations)
            if isinstance(parsed, list):
                return parsed
        except:
            pass
        recommendations_list = []
        current_rec = {}
        lines = recommendations.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith('Title:'):
                if current_rec:
                    recommendations_list.append(current_rec)
                current_rec = {'title': line[6:].strip()}
            elif line.startswith('Description:'):
                current_rec['description'] = line[12:].strip()
            elif line.startswith('Why it matters:'):
                current_rec['why_matters'] = line[15:].strip()
            elif line.startswith('Recommendation:'):
                current_rec['recommendation'] = line[15:].strip()
        if current_rec:
            recommendations_list.append(current_rec)
        return recommendations_list
    parsed_recommendations = parse_recommendations(recommendations)
    # Build results_matrix for the matrix table
    results_matrix = {}
    for response in responses:
        capability_id = response[0]
        lens_id = response[1]
        answer = response[2]
        score = response[3]
        improvement_suggestions = response[4]
        if capability_id not in results_matrix:
            results_matrix[capability_id] = {}
        results_matrix[capability_id][lens_id] = {
            'answer': answer,
            'score': score,
            'improvement_suggestions': improvement_suggestions
        }
    conn.close()
    html_content = render_template('pdf_report.html',
                                 domain=domain,
                                 status=status,
                                 overall_percentage=correct_overall_percentage,
                                 parsed_recommendations=parsed_recommendations,
                                 responses=responses,
                                 created_at=created_at,
                                 updated_at=updated_at,
                                 total_possible_points=total_possible_points,
                                 total_questions=total_questions,
                                 total_score=total_score,
                                 industry_avg=industry_avg,
                                 user_score=user_score,
                                 has_benchmark_data=has_benchmark_data,
                                 results_matrix=results_matrix,
                                 capabilities=domain_capabilities,
                                 lenses=lenses,
                                 domain_benchmarks=domain_benchmarks,
                                 domain_scores=domain_scores,
                                 unique_questions_answered=unique_questions_answered,
                                 raw_average=raw_average,
                                 user_maturity_label=user_maturity_label
                                 )
    try:
        from weasyprint import HTML
        pdf = HTML(string=html_content).write_pdf()
        from io import BytesIO
        pdf_buffer = BytesIO(pdf) if pdf else BytesIO()
        pdf_buffer.seek(0)
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f'finops_assessment_{assessment_id}.pdf',
            mimetype='application/pdf'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF: {str(e)}'}), 500

@utils_bp.route('/export_xlsx/<int:assessment_id>')
def export_xlsx(assessment_id):
    """Export assessment results matrix as XLSX"""
    if 'user_id' not in session:
        return redirect(url_for('auth.index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get assessment details
    cursor.execute('''
        SELECT scope_id, domain, status, overall_percentage, recommendations, created_at, updated_at
        FROM assessments 
        WHERE id = ? AND user_id = ?
    ''', (assessment_id, session['user_id']))
    assessment = cursor.fetchone()
    
    if not assessment:
        conn.close()
        return render_template('dashboard.html', error='Assessment not found')
    
    scope_id, domain, status, overall_percentage, recommendations, created_at, updated_at = assessment
    
    # Get responses (only the latest response per capability_id + lens_id combination)
    cursor.execute('''
        SELECT capability_id, lens_id, answer, score, improvement_suggestions
        FROM responses r1
        WHERE assessment_id = ? 
        AND id = (
            SELECT MAX(id) 
            FROM responses r2 
            WHERE r2.assessment_id = r1.assessment_id 
            AND r2.capability_id = r1.capability_id 
            AND r2.lens_id = r1.lens_id
        )
        ORDER BY capability_id, lens_id
    ''', (assessment_id,))
    responses = cursor.fetchall()
    
    # Get capabilities and lenses
    if domain == 'Complete Assessment':
        from data.capabilities import CAPABILITIES, LENSES
        domain_capabilities = CAPABILITIES
        lenses = LENSES
    else:
        from data.capabilities import CAPABILITIES, LENSES
        domain_capabilities = [cap for cap in CAPABILITIES if cap['domain'] == domain]
        lenses = LENSES
    
    # Build results matrix
    results_matrix = {}
    for response in responses:
        capability_id = response[0]
        lens_id = response[1]
        answer = response[2]
        score = response[3]
        improvement_suggestions = response[4]
        if capability_id not in results_matrix:
            results_matrix[capability_id] = {}
        results_matrix[capability_id][lens_id] = {
            'answer': answer,
            'score': score,
            'improvement_suggestions': improvement_suggestions
        }
    
    conn.close()
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Results Matrix"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        domain_header_font = Font(bold=True, color="FFFFFF")
        domain_header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        capability_font = Font(bold=True)
        capability_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Score cell styles
        score_styles = {
            0: PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
            1: PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            2: PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
            3: PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
            4: PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
        }
        
        # Add title
        ws['A1'] = f"FinOps Assessment Results Matrix - {domain}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Add assessment info
        ws['A3'] = f"Domain: {domain}"
        ws['A4'] = f"Status: {status}"
        ws['A5'] = f"Generated: {updated_at[:10] if updated_at else 'N/A'}"
        
        # Add matrix header
        row = 8
        headers = ['Capability', 'Knowledge (30%)', 'Process (25%)', 'Metrics (20%)', 'Adoption (20%)', 'Automation (5%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        row = 9
        if domain == 'Complete Assessment':
            # Group by domain for complete assessment
            from itertools import groupby
            for domain_name, domain_capabilities_group in groupby(domain_capabilities, key=lambda x: x['domain']):
                # Add domain header
                domain_cell = ws.cell(row=row, column=1, value=domain_name)
                domain_cell.font = domain_header_font
                domain_cell.fill = domain_header_fill
                domain_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Add capabilities for this domain
                for capability in domain_capabilities_group:
                    # Capability name
                    cap_cell = ws.cell(row=row, column=1, value=capability['name'])
                    cap_cell.font = capability_font
                    cap_cell.fill = capability_fill
                    
                    # Scores for each lens
                    for col, lens in enumerate(lenses, 2):
                        score = results_matrix.get(capability['id'], {}).get(lens['id'], {}).get('score', '-')
                        cell = ws.cell(row=row, column=col, value=score)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Apply color based on score
                        if isinstance(score, int) and score in score_styles:
                            cell.fill = score_styles[score]
                    
                    row += 1
        else:
            # Single domain assessment
            for capability in domain_capabilities:
                # Capability name
                cap_cell = ws.cell(row=row, column=1, value=capability['name'])
                cap_cell.font = capability_font
                cap_cell.fill = capability_fill
                
                # Scores for each lens
                for col, lens in enumerate(lenses, 2):
                    score = results_matrix.get(capability['id'], {}).get(lens['id'], {}).get('score', '-')
                    cell = ws.cell(row=row, column=col, value=score)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply color based on score
                    if isinstance(score, int) and score in score_styles:
                        cell.fill = score_styles[score]
                
                row += 1
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Calculate totals
        total_score = sum(response[3] for response in responses) if responses else 0
        total_questions = len(responses) if responses else 0
        total_possible_points = total_questions * 4 if total_questions > 0 else 0
        overall_percentage = (total_score / total_possible_points) * 100 if total_possible_points > 0 else 0
        raw_average = (total_score / total_questions) if total_questions > 0 else 0
        
        # Add summary data
        summary_ws['A1'] = "Assessment Summary"
        summary_ws['A1'].font = Font(size=16, bold=True)
        
        summary_ws['A3'] = "Metric"
        summary_ws['B3'] = "Value"
        summary_ws['A3'].font = Font(bold=True)
        summary_ws['B3'].font = Font(bold=True)
        
        summary_data = [
            ("Domain", domain),
            ("Status", status),
            ("Questions Answered", total_questions),
            ("Total Score", total_score),
            ("Total Possible Points", total_possible_points),
            ("Overall Maturity (%)", f"{overall_percentage:.1f}%"),
            ("Average Score (0-4)", f"{raw_average:.1f}"),
            ("Generated", updated_at[:10] if updated_at else 'N/A')
        ]
        
        for i, (metric, value) in enumerate(summary_data, 4):
            summary_ws[f'A{i}'] = metric
            summary_ws[f'B{i}'] = value
        
        # Auto-adjust column widths
        for ws in [wb.active, summary_ws]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f'finops_results_matrix_{assessment_id}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

# Utility functions
def markdown_filter(text):
    """Convert markdown to HTML"""
    if not text:
        return ""
    
    # Fix URL links
    def fix_url_links(md):
        # Replace [https://...](https://...) with just the URL as a clickable link
        import re
        pattern = r'\[(https?://[^\]]+)\]\(https?://[^)]+\)'
        return re.sub(pattern, r'\1', md)
    
    # Convert markdown to HTML
    html_text = markdown.markdown(fix_url_links(text), extensions=['fenced_code', 'tables'])
    
    # Fix real example links
    def fix_real_example_links(html_text):
        # Replace links that point to finops.org
        import re
        pattern = r'<a href="https://www\.finops\.org[^"]*">([^<]+)</a>'
        return re.sub(pattern, r'\1', html_text)
    
    return fix_real_example_links(html_text) 